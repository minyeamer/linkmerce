from __future__ import annotations

from linkmerce.common.transform import ExcelTransformer, DuckDBTransformer

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    from typing import Literal


class RocketSettlement(DuckDBTransformer):
    """쿠팡 로켓그로스 정산현황의 정산 리포트 목록을 변환 및 적재하는 클래스.

    - **Extractor**: `RocketSettlement`

    - **Parser** ( *parser_class: input_type -> output_type* ):
        `JsonTransformer: dict -> list[dict]`

    - **Table** ( *table_key: table_name* ):
        `table: coupang_rocket_settlement`

    Parameters
    ----------
    **NOTE** DuckDB 쿼리 실행에 필요한 파라미터를 `transform` 메서드 호출 시 함께 전달해야 한다.

    vendor_id: str
        업체 코드
    """

    extractor = "RocketSettlement"
    tables = {"table": "coupang_rocket_settlement"}
    parser = "json"
    parser_config = dict(
        dtype = dict,
        scope = "settlementStatusReports",
        fields = {
            ".": [
                "settlementGroupKey", "settlementRatio", "finalSettlementAmount",
                "settlementPeriodStartDate", "settlementPeriodEndDate"
            ],
            "settlementStatusReportDetail": [
                "totalSalesAmount", "totalRefundedAmount", "totalTakeRateAmountWithVat", "totalSellerDiscount",
                "totalSellerFundedInstantDiscount", "totalSellerFundedDownloadDiscount", "totalPayableAmount",
                "totalMilkRunDeductionAmount", "totalAdSalesDeductionAmount", "totalAdditionalDeductionAmount",
                "totalNegativeDeductionAmount", "totalFinalCfsFeeDeductionAmount", "totalWarehousingFeeDeductionAmount",
                "totalFulfillmentFeeDeductionAmount", "totalStorageFeeDeductionAmount",
                "totalCreturnReverseShippingFeeDeductionAmount", "totalCreturnGradingFeeDeductionAmount",
                "totalVreturnHandlingFeeDeductionAmount", "totalBarcodeLabelingFeeDeductionAmount",
                "totalLastSettlementUnpaidCfsDeductionAmount", "totalPastCfsDeductionAmount",
                "totalCarryOverSettlementDeductionAmount", "totalCfsInventoryCompensationAmount"
            ]
        },
    )
    params = {"vendor_id": "$vendor_id"}


class RocketSalesParser(ExcelTransformer):
    """쿠팡 로켓그로스 정산현황의 '판매 수수료 리포트'를 파싱하는 클래스."""

    header = 2
    fields = [
        "주문ID", "등록상품 ID", "옵션ID", "SKU ID", "등록상품명", "옵션명", "카테고리ID", "카테고리명",
        "거래유형", "정산유형", "판매가(A)", "판매수량(B)", "판매액(A*B)", "쿠팡지원할인(C)", "매출금액(A*B-C)",
        "즉시할인쿠폰(D)", "다운로드쿠폰(E)", "판매자할인쿠폰(D+E)", "정산대상액", "판매수수료", "판매수수료 VAT",
        "매출인식일", "정산주기(종료일)"
    ]


class RocketShippingParser(ExcelTransformer):
    """쿠팡 로켓그로스 정산현황의 '입출고비/배송비 리포트'를 파싱하는 클래스."""

    header = None
    fields = [
        "주문ID", "배송ID", "등록상품 ID", "옵션ID", "SKU ID", "등록상품명", "옵션명", "1차", "2차",
        "개별포장 상품 사이즈", "물류센터", "거래유형", "정산유형", "판매수량", "발생비용(A)", "할인가(B)",
        {"추가비용": None}, "주문일", "매출인식일", "정산주기(종료일)"
    ]

    def parse(self, obj: bytes, **kwargs) -> list[dict]:
        """`입출고비`, `배송비` 시트별로 각각 데이터를 읽고 동일한 스키마로 병합해 반환한다."""
        from linkmerce.utils.excel import filter_warnings
        from io import BytesIO
        import openpyxl
        filter_warnings()

        wb = openpyxl.load_workbook(BytesIO(obj))
        report = list()

        for sheet_name in ["입출고비", "배송비"]:
            ws = wb[sheet_name]
            headers1 = [cell.value for cell in next(ws.iter_rows(min_row=7, max_row=7))]
            headers2 = [cell.value for cell in next(ws.iter_rows(min_row=8, max_row=8))]
            headers = [(header2 if header2 else header1) for header1, header2 in zip(headers1, headers2)]
            report += [dict(zip(headers, row)) for row in ws.iter_rows(min_row=9, values_only=True)]

        return report


class RocketSettlementDownload(DuckDBTransformer):
    """쿠팡 로켓그로스 정산현황의 정산 리포트를 변환 및 적재하는 클래스.

    - **Extractor**: `RocketSettlementDownload`

    - **Parsers** ( *parser_class: input_type -> output_type* ):
        1. `RocketSalesParser: bytes -> list[dict]`
        2. `RocketShippingParser: bytes -> list[dict]`

    - **Tables** ( *table_key: table_name (description)* ):
        1. `sales: coupang_rocket_sales` (판매 수수료 리포트)
        2. `shipping: coupang_rocket_shipping` (입출고비/배송비 리포트)

    Parameters
    ----------
    **NOTE** 입력 데이터 유형을 특정하기 위한 파라미터를 `transform` 메서드 호출 시 함께 전달해야 한다.

    report_type: str
        정산 리포트 유형
            - `CATEGORY_TR`: 판매 수수료 리포트
            - `WAREHOUSING_SHIPPING`: 입출고비/배송비 리포트

    **NOTE** DuckDB 쿼리 실행에 필요한 파라미터를 `transform` 메서드 호출 시 함께 전달해야 한다.

    vendor_id: str
        업체 코드
    """

    extractor = "RocketSettlementDownload"
    queries = ["create", "bulk_insert_sales", "bulk_insert_shipping"]
    tables = {"sales": "coupang_rocket_sales", "shipping": "coupang_rocket_shipping"}
    params = {"vendor_id": "$vendor_id"}

    def parse(self, obj: bytes, report_type: Literal["CATEGORY_TR", "WAREHOUSING_SHIPPING"], **kwargs) -> list[dict]:
        """`report_type`에 따라 파서를 선택해 초기화 및 실행한다."""
        config = self.parser_config or dict()
        if report_type == "CATEGORY_TR":
            return RocketSalesParser(**config).transform(obj, **kwargs)
        elif report_type == "WAREHOUSING_SHIPPING":
            return RocketShippingParser(**config).transform(obj, **kwargs)
        else:
            self.raise_parse_error(f"Parsing for report type '{report_type}' is not supported.")

    def bulk_insert(self, result: list[dict], report_type: Literal["CATEGORY_TR", "WAREHOUSING_SHIPPING"], **kwargs) -> list:
        """`report_type`에 따라 대상 테이블과 삽입 쿼리를 선택해 실행한다."""
        if len(result) > 0:
            table = "sales" if report_type == "CATEGORY_TR" else "shipping"
            render = {table: self.tables[table], f"{table}_rows": self.expr_rows(f"{table}_rows")}
            query = self.prepare_query(f"bulk_insert_{table}", render=render)
            params = self.render_params(kwargs=kwargs) | {f"{table}_rows": result}
            return self.execute(query, params)
